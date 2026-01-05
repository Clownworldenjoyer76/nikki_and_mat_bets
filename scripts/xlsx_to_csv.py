#!/usr/bin/env python3
import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "docs" / "data"

def die(msg, code=78):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(code)

def xlsx_to_csv(xlsx_path: Path, csv_path: Path):
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)) for v in row])

    if not rows:
        die(f"{xlsx_path} appears to be empty")

    headers = rows[0]
    data = rows[1:]

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data)

def main():
    if len(sys.argv) != 4:
        die("Usage: xlsx_to_csv.py <folder: scores|picks|final> <week-number 1..23> <season>")

    folder = sys.argv[1].strip().lower()
    try:
        week = int(sys.argv[2])
        season = int(sys.argv[3])
    except:
        die("Week and season must be integers")

    if folder not in {"scores", "picks", "final"}:
        die(f"Folder must be one of: scores|picks|final (got '{folder}')")

    if not (1 <= week <= 23):
        die("Week must be 1..23")

    wk_tag = f"wk{week:02d}"
    suffix = folder

    base = DATA_DIR / folder
    xlsx_path = base / f"{season}_{wk_tag}_{suffix}.xlsx"

    if not xlsx_path.exists():
        die(f"Missing XLSX file: {xlsx_path}")

    csv_path = xlsx_path.with_suffix(".csv")

    xlsx_to_csv(xlsx_path, csv_path)

    print(f"Converted XLSX -> CSV")
    print(f"  XLSX: {xlsx_path.relative_to(ROOT)}")
    print(f"  CSV : {csv_path.relative_to(ROOT)}")

if __name__ == "__main__":
    main()
